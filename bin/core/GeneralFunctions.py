# @Author  : Xavier Faure
# @Email   : xavierf@kth.se

import os
import eplus.GeomScripts as GeomScripts
import eplus.Set_Outputs as Set_Outputs
import eplus.Sim_param as Sim_param
import eplus.Load_and_occupancy as Load_and_occupancy
import eplus.DomesticHotWater as DomesticHotWater
import building_geometry.MUBES_pygeoj as MUBES_pygeoj
import eplus.build_fmus as BuildFMUs
import outputs.output_utilities as Utilities
from openpyxl import load_workbook
import openturns as ot
import shutil
import pickle
import pyproj
import numpy as np
import json

def appendBuildCase(StudiedCase,keypath,nbcase,DataBaseInput,MainPath,LogFile,PlotOnly = False, DebugMode = False):
    StudiedCase.addBuilding('Building'+str(nbcase),DataBaseInput,nbcase,MainPath,keypath,LogFile,PlotOnly, DebugMode)
    idf = StudiedCase.building[-1]['BuildIDF']
    building = StudiedCase.building[-1]['BuildData']
    return idf, building

def setSimLevel(idf,building):
    ####################################################################
    #Simulation Level
    #####################################################################
    Sim_param.Location_and_weather(idf,building)
    Sim_param.setSimparam(idf,building)

def setBuildingLevel(idf,building,LogFile,CorePerim = False,FloorZoning = False,ForPlots = False,DebugMode = False):
    ######################################################################################
    #Building Level
    ######################################################################################
    #this is the function that requires the longest time
    GeomScripts.createBuilding(LogFile,idf,building, perim = CorePerim,FloorZoning = FloorZoning,ForPlots=ForPlots,DebugMode = DebugMode)

def setEnvelopeLevel(idf,building):
    ######################################################################################
    #Envelope Level (within the building level)
    ######################################################################################
    #the other geometric element are thus here
    GeomScripts.createRapidGeomElem(idf, building)

def setZoneLevel(idf,building,FloorZoning = False):
    ######################################################################################
    #Zone level
    ######################################################################################
    #control command related equipment, loads and leaks for each zones
    Load_and_occupancy.CreateZoneLoadAndCtrl(idf,building,FloorZoning)

def setExtraEnergyLoad(idf,building):
    if building.DHWInfos:
        DomesticHotWater.createWaterEqpt(idf,building)

def setOutputLevel(idf,building,MainPath,EMSOutputs,OutputsFile):
    #ouputs definitions
    Set_Outputs.AddOutputs(idf,building,MainPath,EMSOutputs,OutputsFile)

def readPathfile(Pathways):
    keyPath = {'epluspath': '', 'Buildingsfile': '', 'Shadingsfile': '','pythonpath': '','GeojsonProperties':''} #these keys are hard written as there as used aftewrad at several places, not all of them are used at the moment
    with open(Pathways, 'r') as PathFile:
        Paths = PathFile.readlines()
        for line in Paths:
            for key in keyPath:
                if key in line:
                    keyPath[key] = os.path.normcase(line[line.find(':') + 1:-1])
    return keyPath

def ReadGeoJsonFile(keyPath,CoordSys = '',toBuildPool = False):
    #print('Reading Input files,...')
    try:
        BuildObjectDict = ReadGeojsonKeyNames(keyPath['GeojsonProperties'])
        Buildingsfile = MUBES_pygeoj.load(keyPath['Buildingsfile'])
        if not toBuildPool: Buildingsfile = checkRefCoordinates(Buildingsfile,CoordSys)
        Shadingsfile = getShadowingFile(keyPath['Buildingsfile'],CoordSys)
        #if not toBuildPool: Shadingsfile = checkRefCoordinates(Shadingsfile)
        return {'BuildObjDict':BuildObjectDict,'Build' :Buildingsfile, 'Shades': Shadingsfile}
    except:
        Buildingsfile = MUBES_pygeoj.load(keyPath['Buildingsfile'])
        if not toBuildPool: Buildingsfile = checkRefCoordinates(Buildingsfile,CoordSys)
        Shadingsfile = getShadowingFile(keyPath['Buildingsfile'],CoordSys)
        #if not toBuildPool: Shadingsfile = checkRefCoordinates(Shadingsfile)
        return {'Build': Buildingsfile, 'Shades': Shadingsfile}

def getShadowingFile(BuildingFilePath,CoordSys):
    Shadingsfile = []
    JSONFile = []
    GeJsonFile = []
    BuildingFileName = os.path.basename(BuildingFilePath)
    JSonTest = os.path.join(os.path.dirname(BuildingFilePath),
                            BuildingFileName[:BuildingFileName.index('.')] + '_Walls.json')
    GeoJsonTest = os.path.join(os.path.dirname(BuildingFilePath), BuildingFileName.replace('Buildings', 'Walls'))
    GeoJsonTest1 = True if 'Walls' in GeoJsonTest else False
    if os.path.isfile(JSonTest):
        JSONFile = JSonTest
    elif os.path.isfile(GeoJsonTest) and GeoJsonTest1:
        GeJsonFile = GeoJsonTest
    else:
        msg = '[Prep. Info] No shadowing wall file found'
    if JSONFile:
        msg = '[Prep. Info] json shadowing walls file found'
        with open(JSONFile) as json_file:
            Shadingsfile = json.load(json_file)
    if GeJsonFile:
        msg = '[Prep. Info] Geojson shadowing walls file found'
        Shadingsfile = MUBES_pygeoj.load(GeJsonFile)
        Shadingsfile = checkRefCoordinates(Shadingsfile,CoordSys)
    print(msg)
    return Shadingsfile

def ListAvailableFiles(keyPath):
    # reading the pathfiles and the geojsonfile
    GlobKey = [keyPath]
    # lets see if the input file is a dir with several geojson files
    multipleFiles = []
    BuildingFiles = ReadGeoJsonDir(GlobKey[0])
    if BuildingFiles:
        if len(BuildingFiles)>1:
            multipleFiles = [FileName[:-8] for FileName in BuildingFiles]
        MainRootPath = GlobKey[0]['Buildingsfile']
        GlobKey[0]['Buildingsfile'] = os.path.join(MainRootPath, BuildingFiles[0])
        for nb, file in enumerate(BuildingFiles[1:]):
            GlobKey.append(GlobKey[-1].copy())
            GlobKey[-1]['Buildingsfile'] = os.path.join(MainRootPath, file)
    return GlobKey, multipleFiles

def ReadGeoJsonDir(keyPath):
    #print('Reading Input dir,...')
    BuildingFiles = []
    if os.path.isdir(keyPath['Buildingsfile']):
        FileList = os.listdir(keyPath['Buildingsfile'])
        for nb,file in enumerate(FileList):
            if file[-8:] == '.geojson':
                #print('Building main input file with file nb: ' + str(nb))
                if not 'Wall' in file:
                    BuildingFiles.append(file)
    return BuildingFiles

def checkRefCoordinates(GeojsonFile,CoordSys):
    if not GeojsonFile:
        return GeojsonFile
    # if 'EPSG' in GeojsonFile.crs['properties']['name']:
    #     return GeojsonFile
    # ##The coordinate system depends on the input file, thus, if specific filter or conversion from one to another,
    # # it should be done here
    if type(CoordSys)==int:
        GeojsonFile = MakeCoordConversion(GeojsonFile, CoordSys)
    #GeojsonFile = MakeCoordConversion(GeojsonFile, CoordSys)
    return GeojsonFile

def MakeCoordConversion(GeojsonFile,CoordSys):
    print('Projecting coordinates of Input file,...')
    transformer = pyproj.Transformer.from_crs("CRS84", "epsg:" + str(
        CoordSys))  # this transformation if done for the France's reference
    for idx, obj in enumerate(GeojsonFile):
        newCoord = []
        for poly in obj.geometry.coordinates:
            newpoly = []
            for vertex in poly:
                newvertex = list(transformer.transform(vertex[0], vertex[1]))
                newpoly.append(
                    newvertex)  # the reversed list and this signe were added after looking at google map and the plot for boston city
            newCoord.append(newpoly)
        obj.geometry.coordinates = newCoord
        obj.geometry.update_centroid()
    return GeojsonFile

def ComputeDistance(v1,v2):
    return ((v2[0]-v1[0])**2+(v2[1]-v1[1])**2)**0.5

def MakePolygonPlots(CaseChoices,Pool2Launch):
    Utilities.makePolyPlots(CaseChoices,Pool2Launch)

def MakeAbsoluteCoord(building,idf = [],roundfactor = 8):
    # we need to change the reference coordinate because precision is needed for boundary conditions definition:
    newfoot = []
    for foot in building.footprint:
        newfoot.append([(round(node[0] + building.RefCoord[0],roundfactor), round(node[1] + building.RefCoord[1],roundfactor)) for node in foot])
    building.footprint = newfoot
    for shade in building.shades.keys():
        newcoord = [(round(node[0] + building.RefCoord[0],roundfactor), round(node[1] + building.RefCoord[1],roundfactor)) for node in
                    building.shades[shade]['Vertex']]
        building.shades[shade]['Vertex'] = newcoord
    new_Agreg = [(round(node[0] + building.RefCoord[0],roundfactor), round(node[1] + building.RefCoord[1],roundfactor)) for node in building.AggregFootprint]
    building.AggregFootprint = new_Agreg
    for Wall in building.AdjacentWalls:
        newcoord = [(round(node[0] + building.RefCoord[0],roundfactor), round(node[1] + building.RefCoord[1],roundfactor)) for node in Wall['geometries']]
        Wall['geometries'] = newcoord
    if idf:
        surfaces = idf.getsurfaces() + idf.getshadingsurfaces() + idf.getsubsurfaces()
        for surf in surfaces:
            for i,node in enumerate(surf.coords):
                try:
                    x,y,z = node[0], node[1], node[2]
                    varx = 'Vertex_' + str(i+1) + '_Xcoordinate'
                    vary = 'Vertex_' + str(i+1) + '_Ycoordinate'
                    varz = 'Vertex_' + str(i+1) + '_Zcoordinate'
                    setattr(surf, varx, round(x + building.RefCoord[0],roundfactor))
                    setattr(surf, vary, round(y + building.RefCoord[1],roundfactor))
                    setattr(surf, varz, round(z,roundfactor))
                except:
                    a=1
        return building, idf
    else:
        return building

def SaveCase(MainPath,SepThreads,CaseName,nbBuild):
    SaveDir = os.path.join(os.path.dirname(os.path.dirname(MainPath)), 'SimResults')
    if not os.path.exists(SaveDir):
        os.mkdir(SaveDir)
    if SepThreads:
        os.rename(os.path.join(MainPath, 'RunningFolder'),
                  os.path.join(SaveDir, CaseName + '_Build_' + str(nbBuild)))
        try:
            os.rename(os.path.join(MainPath,CaseName+'_Logs.log'), os.path.join(os.path.join(SaveDir, CaseName + '_Build_' + str(nbBuild)),CaseName+'_Logs.log'))
        except:
            pass
    else:
        os.rename(os.path.join(os.getcwd(), 'RunningFolder'),
                  os.path.join(SaveDir,CaseName))
        try:
            os.rename(os.path.join(MainPath, CaseName + '_Logs.log'),
                  os.path.join(os.path.join(SaveDir,CaseName), CaseName + '_Logs.log'))
        except:
            pass

def CreateSimDir(CurrentPath,DestinationPath,CaseName,SepThreads,nbBuild,idx,MultipleFile = '',Refresh = False,Verbose = False):
    if not os.path.exists(DestinationPath):
        os.mkdir(os.path.abspath(DestinationPath))
        if Verbose: print(DestinationPath +' folder is created')
    SimDir = os.path.join(os.path.abspath(DestinationPath), CaseName)
    if not os.path.exists(SimDir):
        os.mkdir(SimDir)
        if Verbose: print('[Prep. phase] '+CaseName + ' folder is created')
    elif idx == 0 and Refresh:
        shutil.rmtree(SimDir)
        os.mkdir(SimDir)
        if Verbose: print('[Prep. phase] '+CaseName + ' folder is emptied')
    if SepThreads:
        SimDir = os.path.join(SimDir, 'Build_' + str(nbBuild))
        if not os.path.exists(SimDir):
            os.mkdir(SimDir)
            if Verbose: print('[Prep. phase] '+CaseName +  '/Build_' + str(nbBuild)+' folder is created')
        elif idx == 0 and Refresh:
            shutil.rmtree(SimDir)
            os.mkdir(SimDir)
            if Verbose: print('[Prep. phase] '+CaseName + '/Build_' + str(nbBuild)+' folder is emptied')
    if len(MultipleFile)> 0 :
        SimDir = os.path.normcase(
            os.path.join(SimDir, MultipleFile))
        if not os.path.exists(SimDir):
            os.mkdir(SimDir)
            if Verbose: print('[Prep. phase] '+CaseName + '/'+MultipleFile+ ' folder is created')
        elif idx == 0 and Refresh:
            shutil.rmtree(SimDir)
            os.mkdir(SimDir)
            if Verbose: print('[Prep. phase] '+CaseName + '/' + MultipleFile + ' folder is emptied')
    return SimDir

def getDistType(ParamMethod,Bounds):
    if 'Normal' in ParamMethod:
        return ot.Normal((Bounds[1]+Bounds[0])/2,(Bounds[1]-Bounds[0])/6)
    elif 'Triangular' in ParamMethod:
        return ot.Triangular(Bounds[0],(Bounds[1]+Bounds[0])/2, Bounds[1])
    else:
        #the Uniform law is by default
        return ot.Uniform(Bounds[0],Bounds[1])

def getParamSample(VarName2Change,Bounds,nbruns,ParamMethods):
    # Sampling process if some variable is define in the function's arguments
    # It is currently using the latin hyper cube methods from OpenTURNS package for the sampling generation (LHSExperiment)
    Dist = {}
    LinearVal = {}
    varMethodIdx = {'Idx':[],'Method':[]}
    LinearIdx = 0
    DistIdx = 0
    for idx,param in enumerate(VarName2Change):
        if 'Linear' in ParamMethods[idx]:
            LinearVal[param] = np.linspace(Bounds[idx][0],Bounds[idx][1],nbruns)
            varMethodIdx['Method'].append('Linear')
            varMethodIdx['Idx'].append(LinearIdx)
            LinearIdx+=1
        else:
            Dist[param] = getDistType(ParamMethods[idx],Bounds[idx])
            varMethodIdx['Method'].append('Dist')
            varMethodIdx['Idx'].append(DistIdx)
            DistIdx +=1
    if Dist:
        MakeDist = ot.ComposedDistribution([Dist[x] for x in Dist.keys()])
        OTSample = np.array(ot.LHSExperiment(MakeDist, nbruns).generate())
    if LinearVal:
        LinSample = np.array([[LinearVal[key][x] for key in LinearVal.keys()] for x in range(nbruns)])
    if Dist and LinearVal:
        #both dict need to be implemented but keeping the order of  VarName2change list
        newSample = []
        for idx,mthd in enumerate(varMethodIdx['Method']):
            if idx==0:
                newSample = LinSample[:,varMethodIdx['Idx'][idx]].reshape(nbruns,1) if mthd=='Linear' else \
                                        OTSample[:,varMethodIdx['Idx'][idx]].reshape(nbruns,1)
            else:
                newSample = np.append(newSample,LinSample[:,varMethodIdx['Idx'][idx]].reshape(nbruns,1) if \
                         mthd=='Linear' else OTSample[:,varMethodIdx['Idx'][idx]].reshape(nbruns,1),axis = 1)
        return newSample
    elif Dist:
        return OTSample
    elif LinearVal:
        return LinSample
    return [1]

def CreateFMU(idf,building,nbcase,epluspath,FMUKitPath, SimDir, i,varOut,LogFile,DebugMode):
    print('Building FMU under process...Please wait around 30sec')
    #get the heated zones first and set them into a zonelist
    BuildFMUs.setFMUsINOut(idf, building,varOut)
    idf.saveas('Building_' + str(nbcase) + 'v' + str(i) + '.idf')
    BuildFMUs.buildEplusFMU(FMUKitPath, epluspath, building.WeatherDataFile, os.path.join(SimDir,'Building_' + str(nbcase) + 'v' + str(i) + '.idf'))
    print('FMU created for this building')
    if DebugMode: Write2LogFile('FMU created for this building\n',LogFile)
    if DebugMode: Write2LogFile('##############################################################\n',LogFile)

def ReadGeojsonKeyNames(GeojsonProperties):
    #this is currently not used....
    file = load_workbook(GeojsonProperties).active
    #get the headers
    BldObjName = {}
    for i in range(1,file.max_column+1):
        header = file.cell(row = 1, column = i).internal_value
        if header == 'Field of application':
            AppColNb = i
            for j in range(2,file.max_row+1):
                currentname = file.cell(row = j, column = i).internal_value.replace(' ','_')
                if currentname not in BldObjName.keys():
                    BldObjName[currentname] = {'KeyWord': [],'Attributes':[],'Unit':[],'Format':[],'ExpectedVal':[]}
    # for each objectName, lets get list defined above
    for i in range(1, file.max_column + 1):
        header = file.cell(row=1, column=i).internal_value
        if header == 'Propertie Name':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['KeyWord'].append(file.cell(row = j, column = i).internal_value)
        if header == 'VariableName':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['Attributes'].append(file.cell(row = j, column = i).internal_value)
        if header == 'Unit':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['Unit'].append(file.cell(row = j, column = i).internal_value)
        if header == 'Format':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['Format'].append(file.cell(row = j, column = i).internal_value)
        if header == 'Expected values':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['ExpectedVal'].append(file.cell(row = j, column = i).internal_value)
    return BldObjName

def Write2LogFile(message,LogFile):
    try:
        LogFile.write(message)
    except:
        pass

def ReadZoneOfInterest(ZoneOfInterest,keyWord):
    BldIds = []
    with open(ZoneOfInterest, 'r') as handle:
        FileLines = handle.readlines()
    headers = FileLines[0].split("\t")
    idx = headers.index(keyWord)
    for i,line in enumerate(FileLines[1:]):
        Vals = line.split("\t")
        BldIds.append(Vals[idx])
    return BldIds

def CleanUpLogFiles(MainPath):
    listOfFiles = os.listdir(MainPath)
    file2pick = []
    for file in listOfFiles:
        if '_Logs.log' in file[-9:] :
            file2pick.append(file)
    MainLogFile = open(os.path.join(MainPath, 'AllLogs.log'), 'a')
    for file in file2pick:
        file1 = open(os.path.join(MainPath,file), 'r')
        Lines = file1.readlines()
        file1.close()
        for line in Lines:
            Write2LogFile(line,MainLogFile)
        Write2LogFile('#############################################################\n', MainLogFile)
        os.remove(os.path.join(MainPath,file))
    MainLogFile.close()

def AppendLogFiles(MainPath,BldIDKey):
    file2del = []
    try:
        with open(os.path.join(MainPath, 'AllLogs.log'), 'r') as file:
            Lines = file.readlines()
        file2del = 'AllLogs.log'
    except:
        Liste = os.listdir(MainPath)
        for file in Liste:
            if 'Logs.log' in file:
                with open(os.path.join(MainPath, file), 'r') as extrafile:
                    Lines = extrafile.readlines()
                file2del = file
                break
    if file2del:
        NewLines = []
        flagON = False
        for line in Lines:
            NewLines.append(line)
            if '[Bld ID] '+BldIDKey+' : ' in line:
                flagON = True
                id = line[len('[Bld ID] '+BldIDKey+' : '):-1]
                try:
                    with open(os.path.join(MainPath,'Sim_Results', BldIDKey+'_'+str(id) + '.txt'), 'r') as file:
                        extralines = file.readlines()
                    os.remove(os.path.join(MainPath,'Sim_Results', BldIDKey+'_'+str(id) + '.txt'))
                except:
                    extralines = ['ERROR : No simulations found for this building\n']
            if '[Reported Time]' in line and flagON:
                for extraline in extralines:
                    NewLines.append(extraline)
                flagON = False
        FinalLogFile = open(os.path.join(MainPath, 'FinalLogsCompiled.log'), 'w')
        for line in NewLines:
            Write2LogFile(line, FinalLogFile)
        FinalLogFile.close()
        os.remove(os.path.join(MainPath, file2del))

def setChangedParam(building, ParamVal, VarName2Change, MainPath, DataBaseInput, nbcase, LogFile=[]):
    #there is a loop along the variable name to change and if specific ation are required it should be define here
    # if the variable to change are embedded into several layer of dictionnaries than there is a need to make checks and change accordingly to the correct element
    # here are examples for InternalMass impact using 'InternalMass' keyword in the VarName2Change list to play with the 'WeightperZoneArea' parameter
    # and for ExternalMass impact using 'ExtMass' keyword in the VarName2Change list to play with the 'Thickness' of the wall inertia layer
    roundVal = 3 #this is a more physical based thresehold, so could be 8...."
    for varnum,var in enumerate(VarName2Change):
        if 'InternalMass' in var:
            intmass = building.InternalMass
            intmass['HeatedZoneIntMass']['WeightperZoneArea'] = round(ParamVal[varnum],roundVal)
            setattr(building, var, intmass)
        elif 'ExtMass' in var:
            exttmass = building.Materials
            exttmass['Wall Inertia']['Thickness'] = round(ParamVal[varnum],roundVal)
            setattr(building, var, exttmass)
        elif 'WindowUval' in var:
            building.Materials['Window']['UFactor'] = round(ParamVal[varnum],roundVal)
        elif 'setTempLoL' in var:
            building.setTempLoL = [round(ParamVal[varnum], 3),round(ParamVal[varnum], roundVal)]
        elif 'WallInsuThick' in var:
            exttmass = building.Materials
            exttmass['Wall Insulation']['Thickness'] = max(round(ParamVal[varnum], roundVal),0.005)
            setattr(building, var, exttmass)
        elif 'RoofInsuThick' in var:
            exttmass = building.Materials
            exttmass['Roof Insulation']['Thickness'] = max(round(ParamVal[varnum], roundVal),0.005)
            setattr(building, var, exttmass)
        elif 'MaxShadingDist' in var:
            building.MaxShadingDist = round(ParamVal[varnum], roundVal)
        elif 'IntLoadCurveShape' in var:
            building.IntLoadCurveShape = max(round(ParamVal[varnum], roundVal),1e-6)
            building.IntLoad = building.getIntLoad(MainPath, LogFile)
        elif 'AreaBasedFlowRate' in var:
            building.AreaBasedFlowRate = round(ParamVal[varnum], roundVal)
            building.AreaBasedFlowRateDefault = round(ParamVal[varnum], roundVal)
        else:
            try:
                setattr(building, var, ParamVal[varnum])     #for all other cases with simple float, this line just change the attribute's value directly
            except:
                print('This one needs special care : '+var)

def SetParamSample(SimDir,CaseChoices,SepThreads):
    #the parameter are constructed. the ouput gives a matrix of n parameter to change with nbruns values to simulate
    nbruns = CaseChoices['NbRuns']
    VarName2Change = CaseChoices['VarName2Change']
    Bounds = CaseChoices['Bounds']
    ParamMethods = CaseChoices['ParamMethods']
    if SepThreads:
        Paramfile = os.path.join(SimDir, 'ParamSample.pickle')#os.path.join(os.path.dirname(SimDir), 'ParamSample.pickle')
        if os.path.isfile(Paramfile):
            with open(Paramfile, 'rb') as handle:
                ParamSample = pickle.load(handle)
        else:
            ParamSample = getParamSample(VarName2Change,Bounds,nbruns,ParamMethods)
            if nbruns>1:
                with open(Paramfile, 'wb') as handle:
                    pickle.dump(ParamSample, handle, protocol=pickle.HIGHEST_PROTOCOL)
    else:
        Paramfile = os.path.join(SimDir,'ParamSample.pickle')
        if os.path.isfile(Paramfile):
            with open(Paramfile, 'rb') as handle:
                ParamSample = pickle.load(handle)
        else:
            ParamSample = getParamSample(VarName2Change, Bounds, nbruns,ParamMethods)
            if nbruns > 1:
                with open(Paramfile, 'wb') as handle:
                    pickle.dump(ParamSample, handle, protocol=pickle.HIGHEST_PROTOCOL)
    #lets add a sepcial case for making a sample from posteriors
    if CaseChoices['FromPosteriors']:
        CaseChoices['VarName2Change'] = []
        BldNum = int(SimDir[-SimDir[::-1].index('_'):])
        with open(os.path.join(CaseChoices['PosteriorsDataPath'], 'GlobalMatchedParam.pickle'), 'rb') as handle:
            MatchedData = pickle.load(handle)
        for paramName in MatchedData[CaseChoices['CalibTimeBasis']][BldNum].keys():
            if type(MatchedData[CaseChoices['CalibTimeBasis']][BldNum][paramName]) == np.ndarray:
                CaseChoices['VarName2Change'].append(paramName)
        ParamSample = []
        if CaseChoices['VarName2Change']:
            nbmatches = len(MatchedData[CaseChoices['CalibTimeBasis']][BldNum][CaseChoices['VarName2Change'][0]])
            if nbmatches < 100:
                return ParamSample,CaseChoices
            else:
                for i in range(nbmatches):
                    ParamSet = [float(MatchedData[CaseChoices['CalibTimeBasis']][BldNum][key][i]) for key in CaseChoices['VarName2Change']]
                    if CaseChoices['ECMParam']:
                        if type(CaseChoices['ECMParam'])==list:
                            for paranmidx,pramName in enumerate(CaseChoices['ECMParam']):
                                ParamSet[int(CaseChoices['VarName2Change'].index(pramName))] *= float(
                                    CaseChoices['ECMChange'][paranmidx])
                        else:
                            try: ParamSet[int(CaseChoices['VarName2Change'].index(CaseChoices['ECMParam']))] *= float(CaseChoices['ECMChange'])
                            except: pass
                    ParamSample.append(ParamSet)
                if CaseChoices['ECMParam']:
                    import random
                    random.shuffle(ParamSample)
                    ParamSample = ParamSample[:100]
                ParamSample = np.array(ParamSample)
                CaseChoices['NbRuns'] = len(ParamSample[:,0])
                with open(Paramfile, 'wb') as handle:
                    pickle.dump(ParamSample, handle, protocol=pickle.HIGHEST_PROTOCOL)
    return ParamSample,CaseChoices

def ReadData(line,seperator,header = False):
    Val = [line[:line.index(seperator)]]
    remainline = line[line.index(seperator) + 1:]
    stillhere = 1
    while stillhere == 1:
        try:
            Val.append(remainline[:remainline.index(seperator)])
            remainline = remainline[remainline.index(seperator) + 1:]
        except:
            stillhere = 0
    Val.append(remainline[:-1])
    if header:
        Outputs = {}
        for i in Val:
            Outputs[i] =[]
    else:
        Outputs = Val
    return Outputs

def getInputFile(path,seperator):
    with open(path, 'r') as handle:
        FileLines = handle.readlines()
    Header = ReadData(FileLines[0],seperator,header=True)
    for i,line in enumerate(FileLines[1:]):
            Val = ReadData(line,seperator)
            try: float(Val[0].replace(',','.'))
            except: break
            for id,key in enumerate(Header.keys()):
                Header[key].append(Val[id].replace(',','.'))
    return Header

def ManageGlobalPlots(BldObj,IdfObj,FigCenter,WindSize, PlotBldOnly,nbcase = [],LastBld = False):
    FigCenter.append(BldObj.RefCoord)
    refx = sum([center[0] for center in FigCenter]) / len(FigCenter)
    refy = sum([center[1] for center in FigCenter]) / len(FigCenter)
    FigCentroid = BldObj.RefCoord if PlotBldOnly else (refx, refy)
    # we need to transform the previous relatve coordinates into absolute one in order to make plot of several building keeping their location
    if PlotBldOnly:
        FigCentroid = (0,0)
    else:
        BldObj,IdfObj = MakeAbsoluteCoord(BldObj,IdfObj)
    # compåuting the window size for visualization
    for poly in BldObj.footprint:
        for vertex in poly:
            WindSize = max(ComputeDistance(FigCentroid, vertex), WindSize)
    surf = IdfObj.getsurfaces()
    ok2plot = False
    nbadiab = 0
    adiabsurf = []
    for s in surf:
        if s.Outside_Boundary_Condition == 'adiabatic':
            ok2plot = True
            if s.Name[:s.Name.index('_')] not in adiabsurf:
                adiabsurf.append(s.Name[:s.Name.index('_')])
                nbadiab += 1
    RoofSpecialColor = "firebrick"
    IdfObj.idfname = 'GlobGeoJsonImage'
    IdfObj.view_model(test= True if PlotBldOnly+LastBld>0 else False, FigCenter=FigCentroid, WindSize=2 * WindSize,
                       RoofSpecialColor=RoofSpecialColor)
    return FigCenter,WindSize

if __name__ == '__main__' :
    print('GeneralFunctions.py')