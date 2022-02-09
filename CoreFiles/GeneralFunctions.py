# @Author  : Xavier Faure
# @Email   : xavierf@kth.se

import os
import CoreFiles.GeomScripts as GeomScripts
import CoreFiles.Set_Outputs as Set_Outputs
import CoreFiles.Sim_param as Sim_param
import CoreFiles.Load_and_occupancy as Load_and_occupancy
import CoreFiles.DomesticHotWater as DomesticHotWater
import CoreFiles.MUBES_pygeoj as MUBES_pygeoj
import CoreFiles.BuildFMUs as BuildFMUs
from openpyxl import load_workbook
from SALib.sample import latin
import shutil
import pickle
import pyproj
import numpy as np

def appendBuildCase(StudiedCase,epluspath,nbcase,DataBaseInput,MainPath,LogFile,PlotOnly = False, DebugMode = False):
    StudiedCase.addBuilding('Building'+str(nbcase),DataBaseInput,nbcase,MainPath,epluspath,LogFile,PlotOnly, DebugMode)
    idf = StudiedCase.building[-1]['BuildIDF']
    building = StudiedCase.building[-1]['BuildData']
    return idf, building

def setSimLevel(idf,building):
    ####################################################################
    #Simulation Level
    #####################################################################
    Sim_param.Location_and_weather(idf,building)
    Sim_param.setSimparam(idf,building)

def setBuildingLevel(idf,building,LogFile,CorePerim = False,FloorZoning = False,ForPlots = False,DebugMode = False):
    ######################################################################################
    #Building Level
    ######################################################################################
    #this is the function that requires the longest time
    GeomScripts.createBuilding(LogFile,idf,building, perim = CorePerim,FloorZoning = FloorZoning,ForPlots=ForPlots,DebugMode = DebugMode)


def setEnvelopeLevel(idf,building):
    ######################################################################################
    #Envelope Level (within the building level)
    ######################################################################################
    #the other geometric element are thus here
    GeomScripts.createRapidGeomElem(idf, building)

def setZoneLevel(idf,building,FloorZoning = False):
    ######################################################################################
    #Zone level
    ######################################################################################
    #control command related equipment, loads and leaks for each zones
    Load_and_occupancy.CreateZoneLoadAndCtrl(idf,building,FloorZoning)

def setExtraEnergyLoad(idf,building):
    if building.DHWInfos:
        DomesticHotWater.createWaterEqpt(idf,building)


def setOutputLevel(idf,building,MainPath,EMSOutputs,OutputsFile):
    #ouputs definitions
    Set_Outputs.AddOutputs(idf,building,MainPath,EMSOutputs,OutputsFile)

def readPathfile(Pathways):
    keyPath = {'epluspath': '', 'Buildingsfile': '', 'Shadingsfile': '','pythonpath': '','GeojsonProperties':''} #these keys are hard written as there as used aftewrad at several places
    with open(Pathways, 'r') as PathFile:
        Paths = PathFile.readlines()
        for line in Paths:
            for key in keyPath:
                if key in line:
                    keyPath[key] = os.path.normcase(line[line.find(':') + 1:-1])
    return keyPath

def ReadGeoJsonFile(keyPath):
    #print('Reading Input files,...')
    try:
        BuildObjectDict = ReadGeojsonKeyNames(keyPath['GeojsonProperties'])
        Buildingsfile = MUBES_pygeoj.load(keyPath['Buildingsfile'])
        Shadingsfile = MUBES_pygeoj.load(keyPath['Shadingsfile'])
        Buildingsfile = checkRefCoordinates(Buildingsfile)
        Shadingsfile = checkRefCoordinates(Shadingsfile)
        return {'BuildObjDict':BuildObjectDict,'Build' :Buildingsfile, 'Shades': Shadingsfile}
    except:
        Buildingsfile = MUBES_pygeoj.load(keyPath['Buildingsfile'])
        Shadingsfile = MUBES_pygeoj.load(keyPath['Shadingsfile'])
        Buildingsfile = checkRefCoordinates(Buildingsfile)
        Shadingsfile = checkRefCoordinates(Shadingsfile)
        return {'Build': Buildingsfile, 'Shades': Shadingsfile}

def ReadGeoJsonDir(keyPath):
    #print('Reading Input dir,...')
    BuildingFiles = []
    ShadingWallFiles = []
    if os.path.isdir(keyPath['Buildingsfile']):
        FileList = os.listdir(keyPath['Buildingsfile'])
        for nb,file in enumerate(FileList):
            if 'Buildings' in file:
                #print('Building main input file with file nb: ' + str(nb))
                BuildingFiles.append(file)
            if 'Walls' in file:
                ShadingWallFiles.append(file.replace('Buildings', 'Walls'))

    return BuildingFiles,ShadingWallFiles



def checkRefCoordinates(GeojsonFile):
    if 'EPSG' in GeojsonFile.crs['properties']['name']:
        return GeojsonFile
    ##The coordinate system depends on the input file, thus, if specific filter or conversion from one to another,
    # it should be done here
    else:#if "CRS84" in GeojsonFile.crs['properties']['name']:
        print('Projecting coordinates of Input file,...')
        transformer = pyproj.Transformer.from_crs("CRS84", "epsg:3950") #this transformation if done for the France's reference
        for idx,obj in enumerate(GeojsonFile):
            newCoord = []
            for poly in obj.geometry.coordinates:
                newpoly = []
                for vertex in poly:
                    newpoly.append(list(transformer.transform(vertex[0], vertex[1])))
                newCoord.append(newpoly)
            obj.geometry.coordinates = newCoord
        return GeojsonFile

def ComputeDistance(v1,v2):
    return ((v2[0]-v1[0])**2+(v2[1]-v1[1])**2)**0.5

def MakeAbsoluteCoord(idf,building):
    # we need to convert change the reference coordinate because precision is needed for boundary conditions definition:
    newfoot = []
    for foot in building.footprint:
        newfoot.append([(node[0] + building.RefCoord[0], node[1] + building.RefCoord[1]) for node in foot])
    building.footprint = newfoot
    for shade in building.shades.keys():
        newcoord = [(node[0] + building.RefCoord[0], node[1] + building.RefCoord[1]) for node in
                    building.shades[shade]['Vertex']]
        building.shades[shade]['Vertex'] = newcoord
    newwalls = []
    for Wall in building.AdjacentWalls:
        newcoord = [(node[0] - building.RefCoord[0], node[1] - building.RefCoord[1]) for node in Wall['geometries']]
        Wall['geometries'] = newcoord
    surfaces = idf.getsurfaces() + idf.getshadingsurfaces() + idf.getsubsurfaces()
    for surf in surfaces:
        for i,node in enumerate(surf.coords):
            try:
                x,y,z = node[0], node[1], node[2]
                varx = 'Vertex_' + str(i+1) + '_Xcoordinate'
                vary = 'Vertex_' + str(i+1) + '_Ycoordinate'
                varz = 'Vertex_' + str(i+1) + '_Zcoordinate'
                setattr(surf, varx, x + building.RefCoord[0])
                setattr(surf, vary, y + building.RefCoord[1])
                setattr(surf, varz, z)
            except:
                a=1
    return idf,building



def SaveCase(MainPath,SepThreads,CaseName,nbBuild):
    SaveDir = os.path.join(os.path.dirname(os.path.dirname(MainPath)), 'SimResults')
    if not os.path.exists(SaveDir):
        os.mkdir(SaveDir)
    if SepThreads:
        os.rename(os.path.join(MainPath, 'RunningFolder'),
                  os.path.join(SaveDir, CaseName + '_Build_' + str(nbBuild)))
        try:
            os.rename(os.path.join(MainPath,CaseName+'_Logs.log'), os.path.join(os.path.join(SaveDir, CaseName + '_Build_' + str(nbBuild)),CaseName+'_Logs.log'))
        except:
            pass
    else:
        os.rename(os.path.join(os.getcwd(), 'RunningFolder'),
                  os.path.join(SaveDir,CaseName))
        try:
            os.rename(os.path.join(MainPath, CaseName + '_Logs.log'),
                  os.path.join(os.path.join(SaveDir,CaseName), CaseName + '_Logs.log'))
        except:
            pass

def CreateSimDir(CurrentPath,CaseName,SepThreads,nbBuild,idx,MultipleFile = '',Refresh = False):
    if not os.path.exists(os.path.join(os.path.dirname(os.path.dirname(CurrentPath)),'MUBES_SimResults')):
        os.mkdir(os.path.join(os.path.dirname(os.path.dirname(CurrentPath)),'MUBES_SimResults'))
    SimDir = os.path.normcase(
        os.path.join(os.path.dirname(os.path.dirname(CurrentPath)), os.path.join('MUBES_SimResults', CaseName)))
    if not os.path.exists(SimDir):
        os.mkdir(SimDir)
    elif idx == 0 and Refresh:
        shutil.rmtree(SimDir)
        os.mkdir(SimDir)
    if SepThreads:
        SimDir = os.path.normcase(
            os.path.join(SimDir, 'Build_' + str(nbBuild)))
        if not os.path.exists(SimDir):
            os.mkdir(SimDir)
        elif idx == 0 and Refresh:
            shutil.rmtree(SimDir)
            os.mkdir(SimDir)
    if len(MultipleFile)> 0 :
        SimDir = os.path.normcase(
            os.path.join(SimDir, MultipleFile))
        if not os.path.exists(SimDir):
            os.mkdir(SimDir)
        elif idx == 0 and Refresh:
            shutil.rmtree(SimDir)
            os.mkdir(SimDir)
    return SimDir

def getParamSample(VarName2Change,Bounds,nbruns):
    # Sampling process if someis define int eh function's arguments
    # It is currently using the latin hyper cube methods for the sampling generation (latin.sample)
    Param = [1]
    try:
        if len(VarName2Change) > 0:
            problem = {}
            problem['names'] = VarName2Change
            problem['bounds'] = Bounds  # ,
            problem['num_vars'] = len(VarName2Change)
            # problem = read_param_file(MainPath+'\\liste_param.txt')
            Param = latin.sample(problem, nbruns)
    except: pass
    return Param

def CreatFMU(idf,building,nbcase,epluspath,SimDir, i,varOut,LogFile,DebugMode):
    print('Building FMU under process...Please wait around 30sec')
    #get the heated zones first and set them into a zonelist
    BuildFMUs.setFMUsINOut(idf, building,varOut)
    idf.saveas('Building_' + str(nbcase) + 'v' + str(i) + '.idf')
    BuildFMUs.buildEplusFMU(epluspath, building.WeatherDataFile, os.path.join(SimDir,'Building_' + str(nbcase) + 'v' + str(i) + '.idf'))
    print('FMU created for this building')
    if DebugMode: Write2LogFile('FMU created for this building\n',LogFile)
    if DebugMode: Write2LogFile('##############################################################\n',LogFile)

def ReadGeojsonKeyNames(GeojsonProperties):
    #this is currently ot used....
    file = load_workbook(GeojsonProperties).active
    #get the headers
    BldObjName = {}
    for i in range(1,file.max_column+1):
        header = file.cell(row = 1, column = i).internal_value
        if header == 'Field of application':
            AppColNb = i
            for j in range(2,file.max_row+1):
                currentname = file.cell(row = j, column = i).internal_value.replace(' ','_')
                if currentname not in BldObjName.keys():
                    BldObjName[currentname] = {'KeyWord': [],'Attributes':[],'Unit':[],'Format':[],'ExpectedVal':[]}
    # for each objectName, lets get list defined above
    for i in range(1, file.max_column + 1):
        header = file.cell(row=1, column=i).internal_value
        if header == 'Propertie Name':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['KeyWord'].append(file.cell(row = j, column = i).internal_value)
        if header == 'VariableName':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['Attributes'].append(file.cell(row = j, column = i).internal_value)
        if header == 'Unit':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['Unit'].append(file.cell(row = j, column = i).internal_value)
        if header == 'Format':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['Format'].append(file.cell(row = j, column = i).internal_value)
        if header == 'Expected values':
            for j in range(2, file.max_row + 1):
                BldObjName[file.cell(row=j, column=AppColNb).internal_value.replace(' ','_')]['ExpectedVal'].append(file.cell(row = j, column = i).internal_value)
    return BldObjName

def Write2LogFile(message,LogFile):
    try:
        LogFile.write(message)
    except:
        pass

def ReadZoneOfInterest(ZoneOfInterest,keyWord):
    BldIds = []
    with open(ZoneOfInterest, 'r') as handle:
        FileLines = handle.readlines()
    headers = FileLines[0].split("\t")
    idx = headers.index(keyWord)
    for i,line in enumerate(FileLines[1:]):
        Vals = line.split("\t")
        BldIds.append(Vals[idx])
    return BldIds

def CleanUpLogFiles(MainPath):
    listOfFiles = os.listdir(MainPath)
    file2pick = []
    for file in listOfFiles:
        if '_Logs.log' in file[-9:] :
            file2pick.append(file)
    MainLogFile = open(os.path.join(MainPath, 'AllLogs.log'), 'a')
    for file in file2pick:
        file1 = open(os.path.join(MainPath,file), 'r')
        Lines = file1.readlines()
        file1.close()
        for line in Lines:
            Write2LogFile(line,MainLogFile)
        os.remove(os.path.join(MainPath,file))
    MainLogFile.close()

def AppendLogFiles(MainPath):
    file2del = []
    try:
        with open(os.path.join(MainPath, 'AllLogs.log'), 'r') as file:
            Lines = file.readlines()
        file2del = 'AllLogs.log'
    except:
        Liste = os.listdir(MainPath)
        for file in Liste:
            if 'Logs.log' in file:
                with open(os.path.join(MainPath, file), 'r') as extrafile:
                    Lines = extrafile.readlines()
                file2del = file
                break
    if file2del:
        NewLines = []
        flagON = False
        for line in Lines:
            NewLines.append(line)
            if '[Bld ID] 50A_UUID : ' in line:
                flagON = True
                id = line[20:-1]
                with open(os.path.join(MainPath,'Sim_Results', id + '.txt'), 'r') as file:
                    extralines = file.readlines()
            if '[Reported Time]' in line and flagON:
                for extraline in extralines:
                    NewLines.append(extraline)
                flagON = False
        FinalLogFile = open(os.path.join(MainPath, 'FinalLogsCompiled.log'), 'w')
        for line in NewLines:
            Write2LogFile(line, FinalLogFile)
        FinalLogFile.close()
        os.remove(os.path.join(MainPath, file2del))


def setChangedParam(building,ParamVal,VarName2Change,MainPath,Buildingsfile,Shadingsfile,nbcase,DB_Data,LogFile=[]):
    #there is a loop file along the variable name to change and if specific ation are required it should be define here
    # if the variable to change are embedded into several layer of dictionnaries than there is a need to make checks and change accordingly to the correct element
    # here are examples for InternalMass impact using 'InternalMass' keyword in the VarName2Change list to play with the 'WeightperZoneArea' parameter
    # and for ExternalMass impact using 'ExtMass' keyword in the VarName2Change list to play with the 'Thickness' of the wall inertia layer
    roundVal = 3 #this is a more physical based thresehold, so could be 8...."
    for varnum,var in enumerate(VarName2Change):
        if 'InternalMass' in var:
            intmass = building.InternalMass
            intmass['HeatedZoneIntMass']['WeightperZoneArea'] = round(ParamVal[varnum],roundVal)
            setattr(building, var, intmass)
        elif 'ExtMass' in var:
            exttmass = building.Materials
            exttmass['Wall Inertia']['Thickness'] = round(ParamVal[varnum],roundVal)
            setattr(building, var, exttmass)
        elif 'WindowUval' in var:
            building.Materials['Window']['UFactor'] = round(ParamVal[varnum],roundVal)
        elif 'setTempLoL' in var:
            building.setTempLoL = [round(ParamVal[varnum], 3),round(ParamVal[varnum], roundVal)]
        elif 'WallInsuThick' in var:
            exttmass = building.Materials
            exttmass['Wall Insulation']['Thickness'] = max(round(ParamVal[varnum], roundVal),0.005)
            setattr(building, var, exttmass)
        elif 'RoofInsuThick' in var:
            exttmass = building.Materials
            exttmass['Roof Insulation']['Thickness'] = max(round(ParamVal[varnum], roundVal),0.005)
            setattr(building, var, exttmass)
        elif 'MaxShadingDist' in var:
            building.MaxShadingDist = round(ParamVal[varnum], roundVal)
            building.shades = building.getshade(Buildingsfile[nbcase], Shadingsfile, Buildingsfile,DB_Data.GeomElement,LogFile,PlotOnly = False)
        elif 'IntLoadCurveShape' in var:
            building.IntLoadCurveShape = max(round(ParamVal[varnum], roundVal),1e-6)
            building.IntLoad = building.getIntLoad(MainPath, LogFile)
        elif 'AreaBasedFlowRate' in var:
            building.AreaBasedFlowRate = round(ParamVal[varnum], roundVal)
            building.AreaBasedFlowRateDefault = round(ParamVal[varnum], roundVal)
        else:
            try:
                setattr(building, var, ParamVal[varnum])     #for all other cases with simple float, this line just change the attribute's value directly
            except:
                print('This one needs special care : '+var)

def SetParamSample(SimDir,CaseChoices,SepThreads):
    #the parameter are constructed. the oupute gives a matrix ofn parameter to change with nbruns values to simulate
    nbruns = CaseChoices['NbRuns']
    VarName2Change = CaseChoices['VarName2Change']
    Bounds = CaseChoices['Bounds']
    if SepThreads:
        Paramfile = os.path.join(SimDir, 'ParamSample.pickle')#os.path.join(os.path.dirname(SimDir), 'ParamSample.pickle')
        if os.path.isfile(Paramfile):
            with open(Paramfile, 'rb') as handle:
                ParamSample = pickle.load(handle)
        else:
            ParamSample = getParamSample(VarName2Change,Bounds,nbruns)
            if nbruns>1:
                with open(Paramfile, 'wb') as handle:
                    pickle.dump(ParamSample, handle, protocol=pickle.HIGHEST_PROTOCOL)
    else:
        Paramfile = os.path.join(SimDir,'ParamSample.pickle')
        if os.path.isfile(Paramfile):
            with open(Paramfile, 'rb') as handle:
                ParamSample = pickle.load(handle)
        else:
            ParamSample = getParamSample(VarName2Change, Bounds, nbruns)
            if nbruns > 1:
                with open(Paramfile, 'wb') as handle:
                    pickle.dump(ParamSample, handle, protocol=pickle.HIGHEST_PROTOCOL)
    #lets add a sepcial case for making a sample from posteriors
    if CaseChoices['FromPosteriors']:
        CaseChoices['VarName2Change'] = []
        posteriors = getInputFile(os.path.join(CaseChoices['PosteriorsDataPath'],'yearlybasis_FinalPosteriors_Bld'+SimDir[-2:]+'.csv'), ';')
        for paramName in posteriors.keys():
            CaseChoices['VarName2Change'].append(paramName)
        ParamSample = []
        for i in range(len(posteriors[paramName])):
            ParamSample.append([float(posteriors[key][i]) for key in posteriors.keys()])
        ParamSample = np.array(ParamSample)
        CaseChoices['NbRuns'] = len(ParamSample[:,0])
        with open(Paramfile, 'wb') as handle:
            pickle.dump(ParamSample, handle, protocol=pickle.HIGHEST_PROTOCOL)
    return ParamSample,CaseChoices

def ReadData(line,seperator,header = False):
    Val = [line[:line.index(seperator)]]
    remainline = line[line.index(seperator) + 1:]
    stillhere = 1
    while stillhere == 1:
        try:
            Val.append(remainline[:remainline.index(seperator)])
            remainline = remainline[remainline.index(seperator) + 1:]
        except:
            stillhere = 0
    Val.append(remainline[:-1])
    if header:
        Outputs = {}
        for i in Val:
            Outputs[i] =[]
    else:
        Outputs = Val
    return Outputs

def getInputFile(path,seperator):
    with open(path, 'r') as handle:
        FileLines = handle.readlines()
    Header = ReadData(FileLines[0],seperator,header=True)
    for i,line in enumerate(FileLines[1:]):
            Val = ReadData(line,seperator)
            try: float(Val[0].replace(',','.'))
            except: break
            for id,key in enumerate(Header.keys()):
                Header[key].append(Val[id].replace(',','.'))
    return Header


if __name__ == '__main__' :
    print('GeneralFunctions.py')